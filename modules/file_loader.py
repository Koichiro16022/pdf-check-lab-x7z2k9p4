# modules/file_loader.py
"""
Excelファイル読み込みモジュール
Raw Textで読み込み、型・数式・書式を全て保持
"""

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import hashlib
from io import BytesIO

class ExcelLoader:
    """Excelファイル読み込みクラス"""
    
    def __init__(self, file):
        self.file = file
        self.wb = None
        self.sheet_names = []
        self.file_name = getattr(file, 'name', 'unknown')
        
    def load(self):
        """
        ファイルを読み込む
        data_only=False で数式を保持
        
        Returns:
            tuple: (success: bool, message: str)
        """
        try:
            # ファイルをバイト列として読み込み
            file_bytes = self.file.read()
            self.file.seek(0)  # ファイルポインタをリセット
            
            # openpyxlで読み込み
            self.wb = load_workbook(
                BytesIO(file_bytes),
                data_only=False,  # 数式を保持
                keep_vba=False,   # VBAは無視
                rich_text=True    # リッチテキストを保持
            )
            
            self.sheet_names = self.wb.sheetnames
            
            return True, f"読み込み成功: {len(self.sheet_names)}シート"
        
        except Exception as e:
            return False, f"読み込みエラー: {str(e)}"
    
    def get_sheet(self, sheet_name):
        """
        シートを取得
        
        Args:
            sheet_name: シート名
            
        Returns:
            Worksheet or None
        """
        if sheet_name in self.sheet_names:
            return self.wb[sheet_name]
        return None
    
    #@st.cache_data
    def get_cell_data(_self, ws, row, col):
        """
        セルの全データを取得
        
        Args:
            ws: ワークシート
            row: 行番号（1始まり）
            col: 列番号（1始まり）
        
        Returns:
            dict: セルの全情報
        """
        try:
            cell = ws.cell(row=row, column=col)
            
            # フォント情報
            font_info = {
                'name': cell.font.name if cell.font else None,
                'size': cell.font.size if cell.font else None,
                'bold': cell.font.bold if cell.font else False,
                'italic': cell.font.italic if cell.font else False,
                'underline': cell.font.underline if cell.font else None,
                'color': str(cell.font.color.rgb) if cell.font and cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                'strike': bool(cell.font.strike) if cell.font else False,
                'vert_align': cell.font.vertAlign if cell.font else None,
            }
            
            # 塗りつぶし情報
            fill_info = {
                'type': cell.fill.fill_type if cell.fill else None,
                'start_color': str(cell.fill.start_color.rgb) if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb') else None,
                'end_color': str(cell.fill.end_color.rgb) if cell.fill and cell.fill.end_color and hasattr(cell.fill.end_color, 'rgb') else None
            }
            
            # 罫線情報（スタイル＋色）
            def _border_color(side_obj):
                try:
                    if side_obj and side_obj.color and hasattr(side_obj.color, 'rgb'):
                        rgb_val = side_obj.color.rgb
                        # テーマカラー等でRGB取得不可の場合はNoneを返す
                        # 有効なRGB文字列は6文字(RRGGBB)または8文字(AARRGGBB)
                        if isinstance(rgb_val, str) and len(rgb_val) in (6, 8):
                            return rgb_val
                except Exception:
                    pass
                return None

            border_info = {
                'top': cell.border.top.style if cell.border and cell.border.top else None,
                'bottom': cell.border.bottom.style if cell.border and cell.border.bottom else None,
                'left': cell.border.left.style if cell.border and cell.border.left else None,
                'right': cell.border.right.style if cell.border and cell.border.right else None,
                'top_color': _border_color(cell.border.top) if cell.border else None,
                'bottom_color': _border_color(cell.border.bottom) if cell.border else None,
                'left_color': _border_color(cell.border.left) if cell.border else None,
                'right_color': _border_color(cell.border.right) if cell.border else None,
            }

            # 配置情報（折り返し・縮小・角度を追加）
            alignment_info = {
                'horizontal': cell.alignment.horizontal if cell.alignment else None,
                'vertical': cell.alignment.vertical if cell.alignment else None,
                'wrap_text': cell.alignment.wrap_text if cell.alignment else None,
                'shrink_to_fit': cell.alignment.shrink_to_fit if cell.alignment else None,
                'text_rotation': cell.alignment.text_rotation if cell.alignment else None,
                'indent': cell.alignment.indent if cell.alignment else 0,
            }

            # ハイパーリンク情報
            hyperlink = None
            try:
                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target
            except Exception:
                pass

            # セル保護情報 (No.39 / No.40)
            protection_info = {
                'locked': None,
                'hidden_formula': None,
            }
            try:
                if cell.protection:
                    protection_info['locked'] = cell.protection.locked
                    protection_info['hidden_formula'] = cell.protection.hidden
            except Exception:
                pass

            # コメント情報 (No.101 / No.102)
            comment_info = None
            try:
                if cell.comment:
                    comment_info = {
                        'text': str(cell.comment.text) if cell.comment.text else '',
                        'author': cell.comment.author if cell.comment.author else '',
                    }
            except Exception:
                pass

            return {
                'value': cell.value,
                'data_type': cell.data_type,
                'number_format': cell.number_format,
                'font': font_info,
                'fill': fill_info,
                'border': border_info,
                'alignment': alignment_info,
                'hyperlink': hyperlink,
                'protection': protection_info,
                'comment': comment_info,
                'is_merged': _self._is_merged_cell(ws, row, col),
                'is_hidden': _self._is_hidden(ws, row, col),
                'coordinate': cell.coordinate
            }
        
        except Exception as e:
            # エラー時は空のセルとして扱う
            return {
                'value': None,
                'data_type': 's',
                'number_format': 'General',
                'font': {},
                'fill': {},
                'border': {},
                'alignment': {},
                'hyperlink': None,
                'protection': {},
                'comment': None,
                'is_merged': None,
                'is_hidden': False,
                'coordinate': f"{get_column_letter(col)}{row}",
                'error': str(e)
            }
    
    def _is_merged_cell(self, ws, row, col):
        try:
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                    return str(merged_range)
        except:
            pass
        return None
    
    def _is_hidden(self, ws, row, col):
        try:
            col_letter = get_column_letter(col)
            row_hidden = ws.row_dimensions[row].hidden if row in ws.row_dimensions else False
            col_hidden = ws.column_dimensions[col_letter].hidden if col_letter in ws.column_dimensions else False
            return row_hidden or col_hidden
        except:
            return False
    
    def get_file_hash(self):
        self.file.seek(0)
        file_bytes = self.file.read()
        self.file.seek(0)
        return hashlib.md5(file_bytes).hexdigest()
    
    def get_sheet_info(self, sheet_name):
        ws = self.get_sheet(sheet_name)
        if ws:
            return {
                'name': sheet_name,
                'max_row': ws.max_row,
                'max_column': ws.max_column,
                'dimensions': ws.dimensions
            }
        return None
