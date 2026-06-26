# modules/excel_exporter.py
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from datetime import datetime
from io import BytesIO
import copy


class ExcelExporter:

    def __init__(self):
        self.wb = None
        self.yellow_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )
        self.gray_fill = PatternFill(
            start_color="D3D3D3",
            end_color="D3D3D3",
            fill_type="solid"
        )
        self.orange_fill = PatternFill(
            start_color="FFA500",
            end_color="FFA500",
            fill_type="solid"
        )
        self.red_font = Font(color="FF0000", bold=True)
        self.white_font = Font(color="FFFFFF", bold=True)

    def _write_cell_value(self, new_cell, source_cell):
        """セルの値を書き込む（MergedCellはスキップ）"""
        from openpyxl.cell.cell import MergedCell
        if isinstance(source_cell, MergedCell):
            return  # 結合セルの子セルは読み取り専用のためスキップ

        val = source_cell.value

        if val is None:
            new_cell.value = None
            return

        if isinstance(val, str):
            new_cell.number_format = '@'
            new_cell.value = val
            # '='始まりの数式文字列は文字列型として強制設定
            # （数式として評価されると存在しないシート参照でXML破損の原因になる）
            if val.startswith('='):
                new_cell.data_type = 's'
        else:
            new_cell.value = val

    def create_report_simple(
        self, ws_check, cell_diffs, hidden_diffs, image_diffs, setting_diffs=None
    ):
        self.wb = Workbook()
        self._create_summary_sheet(
            cell_diffs, hidden_diffs, image_diffs, mode="simple"
        )
        self._create_detail_sheet(cell_diffs, hidden_diffs, image_diffs)
        self._create_result_sheet_simple(
            ws_check, cell_diffs, hidden_diffs, image_diffs
        )
        if setting_diffs:
            self._create_settings_diff_sheet(setting_diffs)
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_report_formatted(
        self, ws_check, cell_diffs, hidden_diffs, image_diffs, setting_diffs=None, ws_master=None
    ):
        self.wb = Workbook()
        self._create_summary_sheet(
            cell_diffs, hidden_diffs, image_diffs, mode="formatted"
        )
        self._create_detail_sheet(cell_diffs, hidden_diffs, image_diffs)
        self._create_result_sheet_with_format(
            ws_check, cell_diffs, hidden_diffs, image_diffs, ws_master=ws_master
        )
        if setting_diffs:
            self._create_settings_diff_sheet(setting_diffs)
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_summary_sheet(
        self, cell_diffs, hidden_diffs, image_diffs, mode="simple"
    ):
        ws = self.wb.active
        ws.title = "サマリー"
        mode_text = "書式保持版" if mode == "formatted" else "書式無視版"
        ws['A1'] = f"零(ZERO) 検証結果レポート（{mode_text}）"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)

        row = 4
        ws[f'A{row}'] = "【検証結果サマリー】"
        ws[f'A{row}'].font = Font(size=12, bold=True)

        row += 1
        ws[f'A{row}'] = "不一致セル数:"
        ws[f'B{row}'] = len(cell_diffs)
        ws[f'B{row}'].font = Font(color="FF0000", bold=True)

        row += 1
        ws[f'A{row}'] = "非表示セル差異:"
        ws[f'B{row}'] = len(hidden_diffs)
        ws[f'B{row}'].font = Font(color="FFA500", bold=True)

        row += 1
        ws[f'A{row}'] = "画像差異:"
        ws[f'B{row}'] = len(image_diffs)
        ws[f'B{row}'].font = Font(color="0000FF", bold=True)

        row += 1
        ws[f'A{row}'] = "合計:"
        ws[f'B{row}'] = len(cell_diffs) + len(hidden_diffs) + len(image_diffs)
        ws[f'B{row}'].font = Font(size=12, bold=True)

        row += 2
        ws[f'A{row}'] = "【差異の種類別内訳】"
        ws[f'A{row}'].font = Font(size=12, bold=True)

        diff_types = {}
        for diff in cell_diffs + hidden_diffs:
            for d in diff['differences']:
                dtype = d['type']
                diff_types[dtype] = diff_types.get(dtype, 0) + 1

        row += 1
        ws[f'A{row}'] = "種類"
        ws[f'B{row}'] = "件数"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)

        for dtype, count in sorted(
            diff_types.items(), key=lambda x: x[1], reverse=True
        ):
            row += 1
            ws[f'A{row}'] = dtype
            ws[f'B{row}'] = count

        row += 3
        ws[f'A{row}'] = "【シート構成】"
        ws[f'A{row}'].font = Font(size=12, bold=True)

        row += 1
        ws[f'A{row}'] = "Sheet1: サマリー"
        ws[f'B{row}'] = "検証結果の概要"

        row += 1
        ws[f'A{row}'] = "Sheet2: 不一致詳細リスト"
        ws[f'B{row}'] = "差異の一覧（フィルター機能付き）"

        row += 1
        ws[f'A{row}'] = "Sheet3: 比較結果"
        ws[f'B{row}'] = (
            f"{'書式保持' if mode == 'formatted' else '書式無視'}・色付き結果"
        )
        ws[f'B{row}'].font = Font(color="0000FF", bold=True)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30

    def _create_result_sheet_simple(
        self, ws_check, cell_diffs, hidden_diffs, image_diffs
    ):
        ws = self.wb.create_sheet("比較結果")
        diff_positions = {diff['position']: diff for diff in cell_diffs}
        hidden_positions = {diff['position']: diff for diff in hidden_diffs}

        from openpyxl.cell.cell import MergedCell
        for row in ws_check.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue  # 結合セルの子セルはスキップ
                new_cell = ws.cell(row=cell.row, column=cell.column)
                self._write_cell_value(new_cell, cell)

                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
                    new_cell.number_format = cell.number_format

                position = cell.coordinate

                if position in diff_positions:
                    new_cell.fill = self.yellow_fill
                    new_cell.font = self.red_font
                    diff = diff_positions[position]
                    comment_text = self._create_comment_text(diff)
                    new_cell.comment = Comment(comment_text, "零(ZERO)")
                    new_cell.comment.width = 450
                    new_cell.comment.height = 500

                elif position in hidden_positions:
                    new_cell.fill = self.orange_fill
                    new_cell.font = self.white_font
                    diff = hidden_positions[position]
                    comment_text = (
                        "🟠 非表示セルの差異\n\n"
                        + self._create_comment_text(diff)
                    )
                    new_cell.comment = Comment(comment_text, "零(ZERO)")
                    new_cell.comment.width = 450
                    new_cell.comment.height = 500

                else:
                    new_cell.fill = self.gray_fill

        self._copy_shapes(ws_check, ws)

        for img_diff in image_diffs:
            if img_diff.get('position') and \
               img_diff['position'] not in ['全体', '不明']:
                try:
                    cell = ws[img_diff['position']]
                    cell.fill = self.yellow_fill
                    cell.comment = Comment(
                        f"🖼️ 画像の差異\n\n{img_diff['detail']}",
                        "零(ZERO)"
                    )
                except:
                    pass

    def _create_result_sheet_with_format(
        self, ws_check, cell_diffs, hidden_diffs, image_diffs, ws_master=None
    ):
        ws = self.wb.create_sheet("比較結果")
        diff_positions = {diff['position']: diff for diff in cell_diffs}
        hidden_positions = {diff['position']: diff for diff in hidden_diffs}

        for col_letter, col_dim in ws_check.column_dimensions.items():
            ws.column_dimensions[col_letter].width = col_dim.width
            ws.column_dimensions[col_letter].hidden = col_dim.hidden

        for row_num, row_dim in ws_check.row_dimensions.items():
            ws.row_dimensions[row_num].height = row_dim.height
            ws.row_dimensions[row_num].hidden = row_dim.hidden

        # 差異のある行・列は非表示でも強制的に再表示する
        import re as _re
        _diff_rows = set()
        _diff_cols = set()
        for pos in list(diff_positions.keys()) + list(hidden_positions.keys()):
            _m = _re.match(r'([A-Z]+)(\d+)', pos)
            if _m:
                _diff_cols.add(_m.group(1))
                _diff_rows.add(int(_m.group(2)))
        for _r in _diff_rows:
            ws.row_dimensions[_r].hidden = False
        for _c in _diff_cols:
            ws.column_dimensions[_c].hidden = False

        from openpyxl.cell.cell import MergedCell
        for row in ws_check.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column)
                if isinstance(cell, MergedCell):
                    # 結合セル子は罫線のみコピー（結合前に設定することで外枠罫線を保持）
                    if cell.border:
                        new_cell.border = copy.copy(cell.border)
                    continue
                self._write_cell_value(new_cell, cell)

                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = cell.number_format

                position = cell.coordinate

                if position in diff_positions:
                    new_cell.fill = self.yellow_fill
                    original_font = (
                        copy.copy(cell.font) if cell.font else Font()
                    )
                    new_cell.font = Font(
                        name=original_font.name,
                        size=original_font.size,
                        bold=True,
                        italic=original_font.italic,
                        underline=original_font.underline,
                        color="FF0000"
                    )
                    diff = diff_positions[position]
                    comment_text = self._create_comment_text(diff)
                    new_cell.comment = Comment(comment_text, "零(ZERO)")
                    new_cell.comment.width = 450
                    new_cell.comment.height = 500

                elif position in hidden_positions:
                    new_cell.fill = self.orange_fill
                    original_font = (
                        copy.copy(cell.font) if cell.font else Font()
                    )
                    new_cell.font = Font(
                        name=original_font.name,
                        size=original_font.size,
                        bold=True,
                        italic=original_font.italic,
                        underline=original_font.underline,
                        color="FFFFFF"
                    )
                    diff = hidden_positions[position]
                    comment_text = (
                        "🟠 非表示セルの差異\n\n"
                        + self._create_comment_text(diff)
                    )
                    new_cell.comment = Comment(comment_text, "零(ZERO)")
                    new_cell.comment.width = 450
                    new_cell.comment.height = 500

                else:
                    new_cell.fill = self.gray_fill

        # 結合範囲は罫線コピー後に適用（外枠罫線を正しく保持するため）
        for merged_range in ws_check.merged_cells.ranges:
            ws.merge_cells(str(merged_range))

        self._copy_shapes(ws_check, ws)

        for img_diff in image_diffs:
            if img_diff.get('position') and \
               img_diff['position'] not in ['全体', '不明']:
                try:
                    cell = ws[img_diff['position']]
                    cell.fill = self.yellow_fill
                    cell.comment = Comment(
                        f"🖼️ 画像の差異\n\n{img_diff['detail']}",
                        "零(ZERO)"
                    )
                except:
                    pass

        # 列幅・行高さの差異を比較結果シートにオレンジでハイライト
        if ws_master is not None:
            from openpyxl.utils import column_index_from_string

            def _get_col_widths(w):
                r = {}
                for col, cd in w.column_dimensions.items():
                    v = getattr(cd, 'width', None)
                    if v is not None:
                        r[col] = round(float(v), 2)
                return r

            def _get_row_heights(w):
                r = {}
                for row, rd in w.row_dimensions.items():
                    v = getattr(rd, 'height', None)
                    if v is not None:
                        r[row] = round(float(v), 2)
                return r

            def _col_width_ja(v):
                if v is None:
                    return "初期値"
                return f"{round(v * 0.9375, 1)}"

            cw_m = _get_col_widths(ws_master)
            cw_c = _get_col_widths(ws_check)
            for col in set(list(cw_m.keys()) + list(cw_c.keys())):
                wm, wc = cw_m.get(col), cw_c.get(col)
                if wm != wc:
                    try:
                        col_idx = column_index_from_string(col)
                        target = ws.cell(row=1, column=col_idx)
                        target.fill = self.orange_fill
                        wm_s = _col_width_ja(wm)
                        wc_s = _col_width_ja(wc)
                        target.comment = Comment(
                            f"📐 列幅の差異\n列{col}: 原本 {wm_s} → 比較データ {wc_s}",
                            "零(ZERO)"
                        )
                        target.comment.width = 300
                        target.comment.height = 80
                        ws.row_dimensions[1].hidden = False
                        ws.column_dimensions[col].hidden = False
                    except Exception:
                        pass

            rh_m = _get_row_heights(ws_master)
            rh_c = _get_row_heights(ws_check)
            for row in set(list(rh_m.keys()) + list(rh_c.keys())):
                hm, hc = rh_m.get(row), rh_c.get(row)
                if hm != hc:
                    try:
                        target = ws.cell(row=row, column=1)
                        target.fill = self.orange_fill
                        import math as _math
                        hm_s = f"{_math.floor(hm * 10 + 0.5) / 10}" if hm is not None else "初期値"
                        hc_s = f"{_math.floor(hc * 10 + 0.5) / 10}" if hc is not None else "初期値"
                        target.comment = Comment(
                            f"📐 行高さの差異\n{row}行目: 原本 {hm_s} → 比較データ {hc_s}",
                            "零(ZERO)"
                        )
                        target.comment.width = 300
                        target.comment.height = 80
                        ws.row_dimensions[row].hidden = False
                    except Exception:
                        pass

    def _copy_shapes(self, ws_source, ws_target):
        try:
            if hasattr(ws_source, '_images'):
                for img in ws_source._images:
                    try:
                        new_img = copy.deepcopy(img)
                        ws_target._images.append(new_img)
                    except Exception:
                        pass
        except Exception:
            pass

    def _fix_merged_borders(self, ws_source, ws_target):
        """結合セル範囲の外枠罫線を補正する（右辺・下辺がMergedCellで失われる問題を修正）"""
        from openpyxl.styles import Border
        from openpyxl.cell.cell import MergedCell
        for mr in ws_source.merged_cells.ranges:
            tl_row, tl_col = mr.min_row, mr.min_col
            br_row, br_col = mr.max_row, mr.max_col
            tl_target = ws_target.cell(tl_row, tl_col)
            if isinstance(tl_target, MergedCell):
                continue
            # 右辺の罫線（右端列の各行から取得）
            right_side = None
            for r in range(tl_row, br_row + 1):
                src_cell = ws_source.cell(r, br_col)
                if src_cell.border and src_cell.border.right and src_cell.border.right.style:
                    right_side = copy.copy(src_cell.border.right)
                    break
            # 下辺の罫線（下端行の各列から取得）
            bottom_side = None
            for c_idx in range(tl_col, br_col + 1):
                src_cell = ws_source.cell(br_row, c_idx)
                if src_cell.border and src_cell.border.bottom and src_cell.border.bottom.style:
                    bottom_side = copy.copy(src_cell.border.bottom)
                    break
            if right_side is None and bottom_side is None:
                continue
            current_border = tl_target.border if tl_target.border else Border()
            tl_target.border = Border(
                left=copy.copy(current_border.left),
                right=right_side if right_side else copy.copy(current_border.right),
                top=copy.copy(current_border.top),
                bottom=bottom_side if bottom_side else copy.copy(current_border.bottom),
            )

    def _create_detail_sheet(self, cell_diffs, hidden_diffs, image_diffs):
        ws = self.wb.create_sheet("不一致詳細リスト")

        headers = [
            'No.', 'セル位置', '種類', '差異タイプ', '原本', '比較データ', '詳細'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")

        row = 2

        for i, diff in enumerate(cell_diffs, 1):
            for d in diff['differences']:
                ws.cell(row, 1, i)
                ws.cell(row, 2, diff['position'])
                ws.cell(row, 3, "🟡 セル差異")
                ws.cell(row, 4, d['type'])
                master_str = str(d.get('master', ''))[:100]
                check_str = str(d.get('check', ''))[:100]
                master_cell = ws.cell(row, 5)
                check_cell = ws.cell(row, 6)
                master_cell.number_format = '@'
                check_cell.number_format = '@'
                master_cell.value = master_str
                check_cell.value = check_str
                if master_str.startswith('='):
                    master_cell.data_type = 's'
                if check_str.startswith('='):
                    check_cell.data_type = 's'
                ws.cell(row, 7, d['detail'])
                row += 1

        for i, diff in enumerate(hidden_diffs, 1):
            for d in diff['differences']:
                ws.cell(row, 1, len(cell_diffs) + i)
                ws.cell(row, 2, diff['position'])
                ws.cell(row, 3, "🟠 非表示差異")
                ws.cell(row, 4, d['type'])
                master_str = str(d.get('master', ''))[:100]
                check_str = str(d.get('check', ''))[:100]
                master_cell = ws.cell(row, 5)
                check_cell = ws.cell(row, 6)
                master_cell.number_format = '@'
                check_cell.number_format = '@'
                master_cell.value = master_str
                check_cell.value = check_str
                if master_str.startswith('='):
                    master_cell.data_type = 's'
                if check_str.startswith('='):
                    check_cell.data_type = 's'
                ws.cell(row, 7, d['detail'])
                row += 1

        for i, img_diff in enumerate(image_diffs, 1):
            ws.cell(row, 1, len(cell_diffs) + len(hidden_diffs) + i)
            ws.cell(row, 2, img_diff.get('position', '不明'))
            ws.cell(row, 3, "🖼️ 画像差異")
            ws.cell(row, 4, img_diff['type'])
            ws.cell(row, 5, str(img_diff.get('master', '')))
            ws.cell(row, 6, str(img_diff.get('check', '')))
            ws.cell(row, 7, img_diff['detail'])
            row += 1

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 40

        if row > 1:
            ws.auto_filter.ref = f"A1:G{row-1}"

    def _create_comment_text(self, diff):
        circle_nums = ['①', '②', '③', '④', '⑤', '⑥', '⑦', '⑧', '⑨', '⑩']
        lines = ["🔍 差異の詳細"]

        has_space_only = False
        num_idx = 0

        for d in diff['differences']:
            master_val = str(d.get('master', ''))
            check_val = str(d.get('check', ''))
            num = circle_nums[num_idx] if num_idx < len(circle_nums) else f"({num_idx + 1})"
            num_idx += 1

            if d['type'] == 'value':
                def _add_char_hints(master_text, check_text):
                    """混同しやすいペア(0/O、1/l/I)が両方の値に存在する場合のみアノテーション付与
                    20文字以下: 対象文字を1文字ずつアノテーション
                    21文字以上: 対象文字を含む場合にサマリーを末尾に付与
                    """
                    combined = master_text + check_text
                    confusable_groups = [
                        {'0', 'O', 'o'},  # ゼロ vs オー
                        {'1', 'l', 'I'},  # イチ vs エル vs アイ
                    ]
                    # 数字と英字が両方含まれるグループのみアノテーション対象に追加
                    needs_annotation = set()
                    for group in confusable_groups:
                        present = {c for c in combined if c in group}
                        if any(c.isdigit() for c in present) and any(c.isalpha() for c in present):
                            needs_annotation |= present

                    def _annotate(text):
                        if not needs_annotation:
                            return text
                        if len(text) <= 20:
                            result = ''
                            for c in text:
                                if c in needs_annotation:
                                    if c.isdigit():
                                        result += c + '（数字）'
                                    elif c.isupper():
                                        result += c + '（英大文字）'
                                    else:
                                        result += c + '（英小文字）'
                                else:
                                    result += c
                            return result
                        else:
                            # 21文字以上はサマリー（対象文字が含まれる場合のみ）
                            present_in = {c for c in text if c in needs_annotation}
                            if not present_in:
                                return text
                            categories = []
                            if any(c.isdigit() for c in present_in):
                                categories.append('数字')
                            if any(c.isupper() for c in present_in):
                                categories.append('英大文字')
                            if any(c.islower() for c in present_in):
                                categories.append('英小文字')
                            summary = f'（全て{categories[0]}）' if len(categories) == 1 \
                                      else f'（{"と".join(categories)}の混在）'
                            return text + summary

                    return _annotate(master_text), _annotate(check_text)

                base_m = master_val[:50]
                base_c = check_val[:50]
                if base_m.startswith('=') or base_c.startswith('='):
                    master_display = base_m + '（数式）' if base_m.startswith('=') else base_m
                    check_display  = base_c + '（数式）' if base_c.startswith('=') else base_c
                else:
                    master_display, check_display = _add_char_hints(base_m, base_c)
            elif d['type'] == 'number_format':
                import re
                mv = d.get('master_value')
                cv = d.get('check_value')

                def _fmt_num(value, fmt):
                    from datetime import datetime, date as date_type, time as time_type
                    if value is None:
                        return ''
                    # 時刻型の処理（datetime.time）
                    if isinstance(value, time_type):
                        leading_zero = bool(re.search(r'hh', fmt, re.IGNORECASE))
                        has_seconds  = bool(re.search(r'ss', fmt, re.IGNORECASE))
                        h_str = f"{value.hour:02d}" if leading_zero else str(value.hour)
                        m_str = f"{value.minute:02d}"
                        if has_seconds:
                            return f"{h_str}:{m_str}:{value.second:02d}"
                        return f"{h_str}:{m_str}"
                    # 日付・日時型の処理
                    if isinstance(value, (datetime, date_type)):
                        base_date = f"{value.year}/{value.month}/{value.day}"
                        # ロケール付きフォーマット[$-xxxx]は直接解釈不可→フォールバック
                        if re.search(r'\[\$-', fmt, re.IGNORECASE):
                            if re.search(r'\[\$-(f800|0411)', fmt, re.IGNORECASE):
                                return f"{base_date}（和暦形式）"
                            return base_date
                        # openpyxlの組み込み日付コード→日本語表示へ変換
                        builtin_map = {
                            'mm-dd-yy': 'yyyy/m/d',
                            'd-mmm-yy': 'yyyy/m/d',
                            'd-mmm': 'm/d',
                            'mmm-yy': 'yyyy/m',
                            'm/d/yy': 'yyyy/m/d',
                            'm/d/yy h:mm': 'yyyy/m/d',
                        }
                        f = builtin_map.get(fmt.lower(), fmt.lower())
                        # yyyy/yy/mm/m/dd/d の順に置換（長いものから先に）
                        f = f.replace('yyyy', str(value.year))
                        f = f.replace('yy', str(value.year)[-2:])
                        f = f.replace('mm', f"{value.month:02d}")
                        f = f.replace('m', str(value.month))
                        f = f.replace('dd', f"{value.day:02d}")
                        f = f.replace('d', str(value.day))
                        # 時刻部分は除去
                        f = re.sub(r'\s*h+:mm(:ss)?', '', f).strip()
                        # 置換後もフォーマット記号が残る場合はフォールバック
                        if re.search(r'[\[\\\$]', f):
                            return base_date
                        return f
                    # 数値型の処理
                    if not isinstance(value, (int, float)):
                        return str(value)
                    # ゼロ非表示の検出: フォーマットの第3セクション（ゼロ用）が空なら非表示
                    if value == 0:
                        sections = fmt.split(';')
                        if len(sections) >= 3 and sections[2].strip() == '':
                            return '（非表示）'
                    # 通貨記号の検出（[$¥-411]形式 / "¥"形式 / 直接記号）
                    currency_prefix = ''
                    locale_m = re.match(r'\[\$([^-\]]*)', fmt)
                    if locale_m:
                        currency_prefix = locale_m.group(1)
                    elif fmt.startswith('"'):
                        quote_m = re.match(r'"([^"]*)"', fmt)
                        if quote_m:
                            currency_prefix = quote_m.group(1)
                    elif fmt and fmt[0] in '¥$€£₩':
                        currency_prefix = fmt[0]
                    m = re.search(r'\.([0#]+)', fmt)
                    decimals = len(m.group(1)) if m else 0
                    use_comma = ',' in fmt
                    try:
                        num_str = f"{value:,.{decimals}f}" if use_comma else f"{value:.{decimals}f}"
                        return currency_prefix + num_str
                    except Exception:
                        return str(value)

                master_display = _fmt_num(mv, master_val) if mv is not None else master_val
                check_display = _fmt_num(cv, check_val) if cv is not None else check_val
                def _fmt_code_ja(fmt):
                    """書式コードを日本語に変換"""
                    fl = fmt.lower().strip()
                    if fl in ('general', ''):
                        return '標準'
                    if fl == '@':
                        return '文字列'
                    if re.search(r'0\.0+%|0%', fl):
                        return 'パーセント'
                    if '¥' in fmt or re.search(r'\[\$', fmt):
                        return '通貨'
                    if re.search(r'yyyy|yy|m/d|y/m|mmm', fl):
                        return '日付'
                    if re.search(r'hh?:mm', fl):
                        return '時刻'
                    if re.search(r'^[0#,]+\.?[0#]*$', fl):
                        return '数値'
                    return f'書式（{fmt}）'

                # 表示結果が同じ場合は書式の種類を日本語で表示
                if master_display == check_display:
                    lines.append(f"{num}表示形式が違う:")
                    lines.append(f" 原本: {_fmt_code_ja(master_val)}")
                    lines.append(f" 比較データ: {_fmt_code_ja(check_val)}")
                else:
                    # detail の表示も実際の値に差し替え
                    detail = f"表示形式が違う: {master_display} -> {check_display}"
                    if ': ' in detail:
                        detail_label, detail_value = detail.split(': ', 1)
                        lines.append(f"{num}{detail_label}:")
                        lines.append(f" {detail_value}")
                    else:
                        lines.append(f"{num}{detail}")
                    lines.append(f" 原本: {master_display}")
                    lines.append(f" 比較データ: {check_display}")
                continue
            elif d['type'] == 'data_type':
                mv = d.get('master_value', '')
                cv = d.get('check_value', '')
                type_short = {
                    '数値型': '数値', '文字列型': '文字', '数式型': '数式',
                    '論理値型': '論理値', '日付型': '日付',
                    'エラー型': 'エラー', '不明': '不明'
                }
                reading_hints = {
                    '0': '（ゼロ）', 'O': '（オー）', 'o': '（オー）',
                    '1': '（イチ）', 'l': '（エル）', 'I': '（アイ）'
                }
                m_short = type_short.get(master_val, master_val)
                c_short = type_short.get(check_val, check_val)
                # スペース文字を可視化してから reading_hints を適用
                from datetime import datetime, date as _dt
                def _vis_val(v):
                    if v is None or v == '':
                        return ''
                    # 論理値はExcel表記に合わせてTRUE/FALSEで返す
                    if isinstance(v, bool):
                        return 'TRUE' if v else 'FALSE'
                    if isinstance(v, (datetime, _dt)):
                        return f"{v.year}/{v.month}/{v.day}"
                    return str(v).replace('　', '（全角スペース）').replace(' ', '（半角スペース）')
                mv_vis = _vis_val(mv)
                cv_vis = _vis_val(cv)
                m_hint = reading_hints.get(mv_vis, '')
                c_hint = reading_hints.get(cv_vis, '')
                master_display = f"{m_short}の　{mv_vis}{m_hint}" if mv else master_val
                check_display = f"{c_short}の　{cv_vis}{c_hint}" if cv else check_val
            elif d['type'] in ('fill_color', 'font_color', 'border_top_color',
                               'border_bottom_color', 'border_left_color', 'border_right_color'):
                def _color_ja(v):
                    if v is None or v == '' or v == 'None':
                        return '塗りつぶしあり'
                    if v in ('00000000', '00FFFFFF', '000000'):
                        return '塗りつぶしなし'
                    if v in ('FF000000', '000000'):
                        return '黒'
                    if v in ('FFFFFFFF', 'FFFFFF'):
                        return '白'
                    return f'#{v}'
                master_display = _color_ja(master_val) if master_val else '塗りつぶしなし'
                check_display  = _color_ja(check_val)  if check_val  else '塗りつぶしなし'
            else:
                master_display = master_val[:50]
                check_display = check_val[:50]

            detail = d['detail']
            if ': ' in detail:
                detail_label, detail_value = detail.split(': ', 1)
                lines.append(f"{num}{detail_label}:")
                lines.append(f" {detail_value}")
            else:
                lines.append(f"{num}{detail}")
            lines.append(f" 原本: {master_display}")
            lines.append(f" 比較データ: {check_display}")

            # スペースのみセルのチェック（ラベルを除いた残りが空の場合のみ）
            def _is_space_only(v):
                cleaned = v
                for lbl in ['（半角スペース）', '（全角スペース）',
                             '（特殊スペース・半角幅）', '（特殊スペース・全角幅）',
                             '（見えないスペース）']:
                    cleaned = cleaned.replace(lbl, '')
                return cleaned == '' and v not in ('', '(空白セル)', '(空文字)')
            if _is_space_only(master_val) or _is_space_only(check_val):
                has_space_only = True

        rec_num = circle_nums[num_idx] if num_idx < len(circle_nums) else f"({num_idx + 1})"
        lines.append(f"{rec_num}推奨アクション: 目視で内容を確認してください")

        if has_space_only:
            lines.append("⚠️ 注意: （半角スペース）のみのセルはExcelの仕様上、セル内には表示されません。差異はこのコメントで確認してください。")

        return "\n".join(lines)

    def _create_settings_diff_sheet(self, setting_diffs):
        """シート・ブック設定の差異を別シートに出力"""
        ws = self.wb.create_sheet(title="設定差異")

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        from openpyxl.styles import Alignment, Border, Side
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws['A1'] = "シート・ブック設定 差異レポート"
        ws['A1'].font = Font(size=14, bold=True, color="1F4E79")
        ws.merge_cells('A1:E1')

        ws['A2'] = f"生成日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
        ws['A2'].font = Font(size=9, color="666666")
        ws.merge_cells('A2:E2')

        ws['A3'] = f"検出件数: {len(setting_diffs)} 件"
        ws['A3'].font = Font(size=10, bold=True)
        ws.merge_cells('A3:E3')

        headers = ["No.", "カテゴリ", "項目", "原本", "比較"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        if setting_diffs:
            for row_idx, diff in enumerate(setting_diffs, 6):
                no_cell = ws.cell(row=row_idx, column=1, value=row_idx - 5)
                no_cell.alignment = Alignment(horizontal='center', vertical='center')
                no_cell.border = border

                cat_cell = ws.cell(row=row_idx, column=2, value=diff.get('category', ''))
                cat_cell.fill = category_fill
                cat_cell.alignment = Alignment(horizontal='center', vertical='center')
                cat_cell.border = border

                item_cell = ws.cell(row=row_idx, column=3, value=diff.get('detail', diff.get('type', '')))
                item_cell.alignment = Alignment(vertical='center', wrap_text=True)
                item_cell.border = border

                master_str = str(diff.get('master', ''))
                master_cell = ws.cell(row=row_idx, column=4)
                master_cell.value = master_str
                if master_str.startswith('='):
                    master_cell.data_type = 's'
                master_cell.alignment = Alignment(vertical='center', wrap_text=True)
                master_cell.border = border

                check_str = str(diff.get('check', ''))
                check_cell = ws.cell(row=row_idx, column=5)
                check_cell.value = check_str
                if check_str.startswith('='):
                    check_cell.data_type = 's'
                check_cell.fill = diff_fill
                check_cell.alignment = Alignment(vertical='center', wrap_text=True)
                check_cell.border = border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25

        if setting_diffs:
            ws.auto_filter.ref = f"A5:E{5 + len(setting_diffs)}"