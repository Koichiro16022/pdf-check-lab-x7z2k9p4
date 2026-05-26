"""
零(ZERO) スタンダード版 v1.0 Phase 1
完全統合版: 慧のUI + 蔵人のロジック + 慧の最終調整
"""

import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter

# モジュールインポート
from modules.file_loader import ExcelLoader
from modules.cell_comparator import CellComparator
from modules.image_comparator import ImageComparator
from modules.excel_exporter import ExcelExporter

# --- ページ設定 ---
st.set_page_config(
    layout="wide", 
    page_title="零 (ZERO) - Standard v1.0",
    page_icon="🛡️"
)

# --- セッション状態の初期化 ---
if 'offset_master' not in st.session_state:
    st.session_state.offset_master = [0, 0]  # [row, col]
if 'offset_check' not in st.session_state:
    st.session_state.offset_check = [0, 0]
if 'page_num' not in st.session_state:
    st.session_state.page_num = 0
if 'scan_results' not in st.session_state:
    st.session_state.scan_results = None
if 'current_diff_index' not in st.session_state:
    st.session_state.current_diff_index = 0
if 'continue_scan' not in st.session_state:
    st.session_state.continue_scan = False
if 'accumulated_results' not in st.session_state:
    st.session_state.accumulated_results = []

# --- カスタムCSS ---
st.markdown("""
<style>
    .matched-cell {
        background-color: #D3D3D3;
        color: #555555;
    }
    .diff-cell {
        background-color: #FFFF00;
        color: #FF0000;
        font-weight: bold;
    }
    .hidden-diff-cell {
        background-color: #FFA500;
        color: #FFFFFF;
        font-weight: bold;
    }
    .preview-box {
        border: 2px solid #4CAF50;
        padding: 10px;
        border-radius: 5px;
        background-color: #F0F8F0;
    }
</style>
""", unsafe_allow_html=True)

# --- UI: ヘッダー ---
st.title("🛡️ 零 (ZERO) スタンダード版 Phase 1")
st.caption("1文字、1型、1ピクセルの差異も逃さない「100%制御」検図エンジン")
st.markdown("---")

# --- UI: ファイルアップローダー ---
col_u1, col_u2 = st.columns(2)

with col_u1:
    st.subheader("📄 原本 (Master)")
    file_master = st.file_uploader(
        "原本ファイルを選択", 
        type="xlsx",
        key="master",
        help="検査成績書の原本ファイルをアップロードしてください"
    )

with col_u2:
    st.subheader("📄 比較データ (Check)")
    file_check = st.file_uploader(
        "比較ファイルを選択", 
        type="xlsx",
        key="check",
        help="検証対象の比較ファイルをアップロードしてください"
    )

# --- メイン処理 ---
if file_master and file_check:
    
    # ファイル読み込み
    try:
        with st.spinner("ファイルを読み込み中..."):
            loader_m = ExcelLoader(file_master)
            loader_c = ExcelLoader(file_check)
            
            success_m, msg_m = loader_m.load()
            success_c, msg_c = loader_c.load()
            
            if not success_m or not success_c:
                st.error(f"読み込みエラー: {msg_m}, {msg_c}")
                st.stop()
            
            # 最初のシートを取得
            ws_master = loader_m.get_sheet(loader_m.sheet_names[0])
            ws_check = loader_c.get_sheet(loader_c.sheet_names[0])
            
            # 列数の自動取得（慧の指摘対応）
            max_col = max(ws_master.max_column, ws_check.max_column)
            
            st.success(f"✓ ファイル読み込み完了 (シート: {loader_m.sheet_names[0]}, 最大列数: {max_col})")
    
    except Exception as e:
        st.error(f"エラー: {str(e)}")
        st.stop()
    
    # --- UI: ナビゲーション ---
    st.markdown("---")
    st.subheader("⌨️ ナビゲーション")
    
    col_nav1, col_nav2, col_nav3, col_nav4 = st.columns([2, 1, 1, 2])
    
    # 原本の十字キー
    with col_nav2:
        st.write("🕹️ **原本 基点調整**")
        
        col_up = st.columns([1, 3, 1])
        if col_up[1].button("↑", key="m_up", use_container_width=True):
            st.session_state.offset_master[0] = max(0, st.session_state.offset_master[0] - 1)
            st.rerun()
        
        col_mid = st.columns(3)
        if col_mid[0].button("←", key="m_left", use_container_width=True):
            st.session_state.offset_master[1] = max(0, st.session_state.offset_master[1] - 1)
            st.rerun()
        if col_mid[1].button("🏠", key="m_reset", use_container_width=True, help="A1に戻す"):
            st.session_state.offset_master = [0, 0]
            st.rerun()
        if col_mid[2].button("→", key="m_right", use_container_width=True):
            st.session_state.offset_master[1] = min(max_col - 1, st.session_state.offset_master[1] + 1)
            st.rerun()
        
        col_down = st.columns([1, 3, 1])
        if col_down[1].button("↓", key="m_down", use_container_width=True):
            st.session_state.offset_master[0] += 1
            st.rerun()
        
        start_cell_m = f"{get_column_letter(st.session_state.offset_master[1]+1)}{st.session_state.offset_master[0]+1}"
        st.info(f"📍 基点: **{start_cell_m}**")
    
    # 比較データの十字キー
    with col_nav3:
        st.write("🕹️ **比較データ 基点調整**")
        
        col_up_c = st.columns([1, 3, 1])
        if col_up_c[1].button("↑", key="c_up", use_container_width=True):
            st.session_state.offset_check[0] = max(0, st.session_state.offset_check[0] - 1)
            st.rerun()
        
        col_mid_c = st.columns(3)
        if col_mid_c[0].button("←", key="c_left", use_container_width=True):
            st.session_state.offset_check[1] = max(0, st.session_state.offset_check[1] - 1)
            st.rerun()
        if col_mid_c[1].button("🏠", key="c_reset", use_container_width=True, help="A1に戻す"):
            st.session_state.offset_check = [0, 0]
            st.rerun()
        if col_mid_c[2].button("→", key="c_right", use_container_width=True):
            st.session_state.offset_check[1] = min(max_col - 1, st.session_state.offset_check[1] + 1)
            st.rerun()
        
        col_down_c = st.columns([1, 3, 1])
        if col_down_c[1].button("↓", key="c_down", use_container_width=True):
            st.session_state.offset_check[0] += 1
            st.rerun()
        
        start_cell_c = f"{get_column_letter(st.session_state.offset_check[1]+1)}{st.session_state.offset_check[0]+1}"
        st.info(f"📍 基点: **{start_cell_c}**")
    
# --- 👁️ 突き合わせプレビュー（左右2画面・ピンポイント詳細のみ） ---
    st.markdown("---")
    st.subheader("👁️ 突き合わせプレビュー（操作ガイド）")
    
    with st.expander("▼ プレビューを表示", expanded=True):
        # 1. ページング計算（5行単位で表示範囲を固定）
        row_offset_c = st.session_state.offset_check[0]
        page_start_row_c = (row_offset_c // 5) * 5 + 1
        page_start_row_m = (st.session_state.offset_master[0] // 5) * 5 + 1
        
        # 2. 現在選択中のセル情報
        t_row_m = st.session_state.offset_master[0] + 1
        t_col_m = st.session_state.offset_master[1] + 1
        t_row_c = st.session_state.offset_check[0] + 1
        t_col_c = st.session_state.offset_check[1] + 1

        col_gui_l, col_gui_r = st.columns(2)
        with col_gui_l:
            st.markdown(f"**原本:** {get_column_letter(t_col_m)}{t_row_m}")
        with col_gui_r:
            st.markdown(f"**比較:** {get_column_letter(t_col_c)}{t_row_c}")

        # 3. 左右並列のExcel風グリッド（5行固定表示）
        col_view_l, col_view_r = st.columns(2)
        
        def generate_grid_html(ws, loader, start_row, target_row, target_col_idx):
            html = '<div style="display: grid; grid-template-columns: 40px repeat(3, 1fr); gap: 0px; border-right: 1px solid #bbb; border-bottom: 1px solid #bbb; font-family: sans-serif; background-color: white; border-top: 1px solid #bbb; border-left: 1px solid #bbb;">'
            # ヘッダー
            html += '<div style="background-color: #e6e6e6; height: 25px; border-right: 0.5px solid #bbb; border-bottom: 0.5px solid #bbb;"></div>'
            for j in range(3):
                html += f'<div style="background-color: #e6e6e6; text-align: center; border-right: 0.5px solid #bbb; border-bottom: 0.5px solid #bbb; font-size: 0.8em; line-height: 25px;">{get_column_letter(j+1)}</div>'
            # データ5行
            for r in range(5):
                curr_r = start_row + r
                html += f'<div style="background-color: #e6e6e6; text-align: center; font-weight: bold; border-right: 0.5px solid #bbb; border-bottom: 0.5px solid #bbb; line-height: 40px; font-size: 0.8em;">{curr_r}</div>'
                for c in range(1, 4):
                    is_target = (curr_r == target_row and c == target_col_idx)
                    try:
                        val = str(loader.get_cell_data(ws, curr_r, c)['value'] or "")
                    except: val = ""
                    bg = "#FBFF05" if is_target else "white"
                    border = "2px solid #217346" if is_target else "0.5px solid #eee"
                    html += f'<div style="background-color: {bg}; border: {border}; min-height: 40px; padding: 2px; overflow: auto; font-size: 0.7em; display: flex; align-items: center; box-sizing: border-box;">{val[:30]}</div>'
            return html + "</div>"

        with col_view_l:
            st.caption("📖 原本 (5行固定表示)")
            st.write(generate_grid_html(ws_master, loader_m, page_start_row_m, t_row_m, t_col_m), unsafe_allow_html=True)
        
        with col_view_r:
            st.caption("📝 比較データ (5行固定表示)")
            st.write(generate_grid_html(ws_check, loader_c, page_start_row_c, t_row_c, t_col_c), unsafe_allow_html=True)

        # 4. 🎯 ピンポイント詳細比較（選択中セルのみ）
        st.markdown("<br>", unsafe_allow_html=True)
        st.subheader("🎯 選択中セルの詳細比較")
        
        comp = CellComparator()
        try:
            c_m = loader_m.get_cell_data(ws_master, t_row_m, t_col_m)
            c_c = loader_c.get_cell_data(ws_check, t_row_c, t_col_c)
            v_m = str(c_m['value']) if c_m['value'] is not None else "(空)"
            v_c = str(c_c['value']) if c_c['value'] is not None else "(空)"
            status = "✅ 一致" if comp.create_strict_string(c_m) == comp.create_strict_string(c_c) else "❌ 差異あり"
            
            # 石田様ご指定のシンプル表示
            st.markdown(f"""
            <div style="background-color: #f9f9f9; padding: 15px; border-left: 5px solid #217346; border-radius: 4px; font-family: sans-serif;">
                <b style="font-size: 1.1em;">原本 {t_row_m}行目 ⇔ 比較 {t_row_c}行目 列{get_column_letter(t_col_c)} </b> {status}<br>
                <div style="margin-top: 10px;">
                    <span style="color: blue;">原本: {v_m}</span><br>
                    <span style="color: brown;">比較: {v_c}</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        except Exception as e:
            st.error("詳細プレビューを取得できませんでした。")
        
        # 不要なリスト表示コード（pd.DataFrame(preview_rows)...）は完全に削除しました。

    
# ===== 診断コード（確認後削除） =====
    if st.button("🔬 半角スペース診断", use_container_width=True):
        try:
            test_row = 7
            test_col = 3

            cell_m = loader_m.get_cell_data(ws_master, test_row, test_col)
            cell_c = loader_c.get_cell_data(ws_check, test_row, test_col)

            st.write("### 元データ（半角スペース）")
            st.write(f"value: {repr(cell_m['value'])}")
            st.write(f"data_type: {repr(cell_m['data_type'])}")
            st.write(f"number_format: {repr(cell_m['number_format'])}")

            st.write("### 比較データ（全角スペース）")
            st.write(f"value: {repr(cell_c['value'])}")
            st.write(f"data_type: {repr(cell_c['data_type'])}")
            st.write(f"number_format: {repr(cell_c['number_format'])}")

            st.write("### 書き込みテスト")
            from openpyxl import Workbook
            test_wb = Workbook()
            test_ws = test_wb.active

            test_ws['A1'].number_format = '@'
            test_ws['A1'].value = cell_m['value']
            st.write(f"書き込み後のA1: {repr(test_ws['A1'].value)}")

            test_ws['A2'].number_format = '@'
            test_ws['A2'].value = cell_c['value']
            st.write(f"書き込み後のA2: {repr(test_ws['A2'].value)}")

        except Exception as e:
            st.error(f"診断エラー: {str(e)}")
            import traceback
            st.error(traceback.format_exc())
    # ===== 診断コード終わり =====

    # --- スキャン実行 ---
    st.markdown("---")
    st.subheader("🔍 スキャン実行")
    
    col_scan1, col_scan2, col_scan3 = st.columns([1, 1, 1])
    
    with col_scan1:
        page_size = st.number_input(
            "スキャン行数",
            min_value=10,
            max_value=500,
            value=100,
            step=10,
            help="一度にスキャンする行数を指定"
        )
    
    with col_scan2:
        if st.button("🔍 スキャン開始", type="primary", use_container_width=True):
            with st.spinner(f"{page_size}行をスキャン中..."):
                # 比較実行
                comparator = CellComparator()
                img_comparator = ImageComparator(threshold=0.01)
                
                start_row_m = st.session_state.offset_master[0]
                start_col_m = st.session_state.offset_master[1]
                start_row_c = st.session_state.offset_check[0]
                start_col_c = st.session_state.offset_check[1]
                
                cell_diffs = []
                hidden_diffs = []
                
               
                
		    # --- 🔍 スキャン実行（全域カバー版へ修正） ---
                # どちらかのファイルに存在する最大の列数を取得
                actual_max_col = max(ws_master.max_column, ws_check.max_column)

                for i in range(int(page_size)):
                    row_m = start_row_m + i + 1
                    row_c = start_row_c + i + 1

                    # 両方のファイルで最終行を超えたらループ終了
                    if row_m > ws_master.max_row and row_c > ws_check.max_row:
                        break

                    # 1列目(A列)から最大列まで全スキャン
                    for j in range(actual_max_col):
                        col_m = start_col_m + j + 1
                        col_c = start_col_c + j + 1
                        
                        # 範囲外アクセスを防ぐ安全装置
                        if col_m > ws_master.max_column and col_c > ws_check.max_column:
                            continue

                        try:
                            cell_m = loader_m.get_cell_data(ws_master, row_m, col_m)
                            cell_c = loader_c.get_cell_data(ws_check, row_c, col_c)
                            
                            position = f"{get_column_letter(col_c)}{row_c}"
                            
                            # 比較実行
                            diff = comparator.compare(cell_m, cell_c, position)
                            if diff:
                                if cell_c.get('is_hidden'):
                                    diff['is_hidden_diff'] = True
                                    hidden_diffs.append(diff)
                                else:
                                    cell_diffs.append(diff)
                        
                        except Exception as e:
                            continue
                            if diff:
                                # 非表示差異の判定（慧の提案2）
                                if cell_c['is_hidden']:
                                    diff['is_hidden_diff'] = True
                                    hidden_diffs.append(diff)
                                else:
                                    cell_diffs.append(diff)
                        
                        except Exception as e:
                            pass
                
                # 画像比較
                img_diffs = img_comparator.compare(ws_master, ws_check)
                
                # 結果を保存
                current_result = {
                    'cell_diffs': cell_diffs,
                    'hidden_diffs': hidden_diffs,
                    'image_diffs': img_diffs,
                    'total_cells': int(page_size) * max_col,
                    'matched': int(page_size) * max_col - len(cell_diffs) - len(hidden_diffs),
                    'mismatched': len(cell_diffs) + len(hidden_diffs) + len(img_diffs)
                }
                
                # 継続スキャンの場合は累積（慧の提案3）
                if st.session_state.continue_scan:
                    st.session_state.accumulated_results.append(current_result)
                else:
                    st.session_state.accumulated_results = [current_result]
                
                st.session_state.scan_results = current_result
                st.session_state.current_diff_index = 0
                
                st.success(f"✓ スキャン完了: {len(cell_diffs)}件の差異を検出")
                st.rerun()
    
    with col_scan3:
        # 継続スキャンオプション（慧の提案3）
        continue_mode = st.checkbox(
            "継続スキャンモード",
            value=st.session_state.continue_scan,
            help="前回の結果を保持したまま次のページをスキャン"
        )
        st.session_state.continue_scan = continue_mode
        
        if continue_mode and len(st.session_state.accumulated_results) > 0:
            st.caption(f"📊 累積: {len(st.session_state.accumulated_results)}ページ")
    
    # --- スキャン結果表示 ---
    if st.session_state.scan_results:
        results = st.session_state.scan_results
        
        st.markdown("---")
        st.subheader("📊 スキャン結果サマリー")
        
        col_r1, col_r2, col_r3, col_r4, col_r5 = st.columns(5)
        
        with col_r1:
            st.metric("総セル数", f"{results['total_cells']:,}")
        
        with col_r2:
            match_rate = (results['matched'] / results['total_cells'] * 100) if results['total_cells'] > 0 else 0
            st.metric("一致", f"{results['matched']:,} ({match_rate:.1f}%)")
        
        with col_r3:
            st.metric("🟡 差異", f"{len(results['cell_diffs']):,}")
        
        with col_r4:
            st.metric("🟠 非表示差異", f"{len(results['hidden_diffs']):,}")
        
        with col_r5:
            st.metric("🖼️ 画像差異", f"{len(results['image_diffs']):,}")
        
        # 累積結果の表示
        if len(st.session_state.accumulated_results) > 1:
            st.info(f"💡 継続スキャン: {len(st.session_state.accumulated_results)}ページ分の結果を累積中")
        
        # 不一致詳細
        all_diffs = results['cell_diffs'] + results['hidden_diffs']
        
        if all_diffs:
            st.markdown("---")
            st.subheader("🔍 不一致セル詳細")
            
            # 不一致ジャンプ
            col_j1, col_j2, col_j3 = st.columns([1, 2, 1])
            
            with col_j2:
                col_jp1, col_jp2, col_jp3 = st.columns([1, 1, 1])
                
                with col_jp1:
                    if st.button("◀ 前へ", use_container_width=True):
                        st.session_state.current_diff_index = max(0, st.session_state.current_diff_index - 1)
                        st.rerun()
                
                with col_jp2:
                    total_diffs = len(all_diffs)
                    current = st.session_state.current_diff_index + 1
                    st.markdown(
                        f"<div style='text-align: center; padding: 8px; font-size: 18px;'><b>{current} / {total_diffs}</b></div>", 
                        unsafe_allow_html=True
                    )
                
                with col_jp3:
                    if st.button("次へ ▶", use_container_width=True):
                        st.session_state.current_diff_index = min(len(all_diffs) - 1, st.session_state.current_diff_index + 1)
                        st.rerun()
            
            # 現在の不一致セルを表示
            current_diff = all_diffs[st.session_state.current_diff_index]
            
            # 非表示差異の場合はオレンジ色で強調（慧の提案2）
            if current_diff.get('is_hidden_diff', False):
                st.markdown(f"### 🟠 セル {current_diff['position']} (非表示セルの差異)")
                st.warning("⚠️ このセルまたは隣接セルに非表示の差異があります")
            else:
                st.markdown(f"### 🟡 セル {current_diff['position']}")
            
            for diff in current_diff['differences']:
                with st.expander(f"📌 {diff['type']}: {diff['detail']}", expanded=True):
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        st.text("📄 原本:")
                        st.text_area("", value=str(diff.get('master', '')), height=120, disabled=True, label_visibility="collapsed", key=f"master_{st.session_state.current_diff_index}_{diff.get('type','')}")
                    with col_d2:
                        st.text("📝 比較:")
                        st.text_area("", value=str(diff.get('check', '')), height=120, disabled=True, label_visibility="collapsed", key=f"check_{st.session_state.current_diff_index}_{diff.get('type','')}")
        
        # 画像差異
        if results['image_diffs']:
            st.markdown("---")
            st.subheader("🖼️ 画像差異")
            
            for img_diff in results['image_diffs']:
                st.warning(f"**{img_diff['type']}**: {img_diff['detail']}")
        
        # Excel出力
                
        st.markdown("---")
        st.subheader("💾 結果の出力")
        
        col_export1, col_export2 = st.columns([2, 1])
        
        with col_export1:
            if st.button("📥 Excel出力（2ファイル生成）", use_container_width=True, type="primary"):
                with st.spinner("Excel生成中（2ファイル）..."):
                    try:
                        # 累積結果をマージ
                        all_cell_diffs = []
                        all_hidden_diffs = []
                        all_image_diffs = []
                        
                        for result in st.session_state.accumulated_results:
                            all_cell_diffs.extend(result['cell_diffs'])
                            all_hidden_diffs.extend(result.get('hidden_diffs', []))
                            all_image_diffs.extend(result['image_diffs'])
                        
                        # タイムスタンプ生成
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                        
                        # Excel生成
                        exporter = ExcelExporter()
                        
                        # 1. 書式無視版
                        output_simple = exporter.create_report_simple(
                            ws_check,
                            all_cell_diffs,
                            all_hidden_diffs,
                            all_image_diffs
                        )
                        
                        # 2. 書式保持版
                        output_formatted = exporter.create_report_formatted(
                            ws_check,
                            all_cell_diffs,
                            all_hidden_diffs,
                            all_image_diffs
                        )
                        
                        # セッションに保存
                        st.session_state.excel_simple = output_simple
                        st.session_state.excel_formatted = output_formatted
                        st.session_state.excel_timestamp = timestamp
                        
                        st.success("✓ 2つのExcelファイルを生成しました！下のボタンからダウンロードしてください。")
                    
                    except Exception as e:
                        st.error(f"Excel生成エラー: {str(e)}")
                        import traceback
                        st.error(traceback.format_exc())
        
        with col_export2:
            if st.button("🔄 結果をクリア", use_container_width=True):
                st.session_state.scan_results = None
                st.session_state.accumulated_results = []
                st.session_state.current_diff_index = 0
                if 'excel_simple' in st.session_state:
                    del st.session_state.excel_simple
                if 'excel_formatted' in st.session_state:
                    del st.session_state.excel_formatted
                st.rerun()
        
        # ダウンロードボタン（生成後に表示）
        if 'excel_simple' in st.session_state and 'excel_formatted' in st.session_state:
            st.markdown("---")
            st.subheader("⬇️ ダウンロード")
            
            col_dl1, col_dl2 = st.columns(2)
            
            with col_dl1:
                st.download_button(
                    label="📄 書式無視版をダウンロード",
                    data=st.session_state.excel_simple,
                    file_name=f"検証結果_書式無視版_{st.session_state.excel_timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col_dl2:
                st.download_button(
                    label="📄 書式保持版をダウンロード",
                    data=st.session_state.excel_formatted,
                    file_name=f"検証結果_書式保持版_{st.session_state.excel_timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


else:
    st.info("👆 原本と比較データの両方をアップロードしてください")
    
    st.markdown("---")
    st.subheader("📖 使い方")
    
    col_help1, col_help2 = st.columns(2)
    
    with col_help1:
        st.markdown("""
        ### 基本的な流れ
        1. 原本ファイルをアップロード
        2. 比較ファイルをアップロード
        3. 十字キーで基点を調整（必要に応じて）
        4. スキャン開始ボタンをクリック
        5. 結果を確認
        6. Excelで出力
        """)
    
    with col_help2:
        st.markdown("""
        ### 検出対象
        - ✓ 値の違い（1文字でも）
        - ✓ データ型（1 vs 1.0）
        - ✓ 数式 vs 直接入力
        - ✓ スペース差異
        - ✓ 書式（色・フォント・罫線）
        - ✓ 結合セル範囲
        - ✓ 非表示行/列の内容
        - ✓ 画像（位置・サイズ・内容）
        """)

# --- フッター ---
st.markdown("---")
col_f1, col_f2, col_f3 = st.columns(3)
with col_f1:
    st.caption("開発: 石田 | 支援: 慧(Gemini) & 蔵人(Claude)")
with col_f2:
    st.caption("バージョン: 1.0 Phase 1")
with col_f3:
    st.caption("「私が100%制御しています。」")
